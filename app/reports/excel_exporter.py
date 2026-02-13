# app/reports/excel_exporter.py

import pandas as pd
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ======================================================
# 🔧 EXPORTADOR EXCEL FORMATEADO
# ======================================================

def exportar_excel_formateado(df, archivo, nombre_reporte="Reporte", tipo_formato="general"):
    """
    Crea un Excel con una hoja por tienda, con formato visual profesional.
    
    Args:
        df: DataFrame con los datos a exportar
        archivo: Nombre del archivo de salida
        nombre_reporte: Título del reporte
        tipo_formato: Tipo de formato a aplicar
            - "general": Formato estándar (default)
            - "picking": Formato para picking con header de control
    """

    if df.empty:
        raise ValueError("El DataFrame está vacío, no se puede exportar.")

    # 🔍 Detectar columna de tienda
    if "tienda" in df.columns:
        col_tienda = "tienda"
    elif "tienda_origen" in df.columns:
        col_tienda = "tienda_origen"
    elif "tienda_destino" in df.columns:
        col_tienda = "tienda_destino"
    else:
        raise KeyError(
            f"No se encontró ninguna columna de tienda en el DataFrame.\n"
            f"Columnas disponibles: {list(df.columns)}"
        )

    # --- Preparar datos según tipo de formato ---
    df_procesado = df.copy()
    
    if tipo_formato == "picking":
        column_mapping = {
            'c_barra': 'Cod.Barras',
            'd_marca': 'Marca',
            'color': 'Color',
            'cantidad_a_despachar': 'Cantidad',
            'observacion': 'Observacion'
        }
        
        columnas_picking = [col for col in column_mapping.keys() if col in df_procesado.columns]
        df_procesado = df_procesado[[col_tienda] + columnas_picking].copy()
        df_procesado = df_procesado.rename(columns=column_mapping)

    # --- Crear workbook manualmente ---
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    alt_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    control_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")

    for tienda, df_tienda in df_procesado.groupby(col_tienda, sort=True):
        hoja_nombre = str(tienda)[:25]
        ws = wb.create_sheet(title=hoja_nombre)
        
        # Preparar datos para esta tienda
        if tipo_formato == "picking":
            df_escribir = df_tienda.drop(columns=[col_tienda])
        else:
            df_escribir = df_tienda
        
        # DIFERENCIA CLAVE: crear header ANTES de escribir datos
        if tipo_formato == "picking":
            _crear_hoja_picking(ws, df_escribir, hoja_nombre, nombre_reporte, thin, medium, header_fill, alt_fill, control_fill)
        else:
            _crear_hoja_general(ws, df_escribir, nombre_reporte, thin, header_fill, alt_fill)

    wb.save(archivo)
    print(f"\n🖨️ Archivo listo y formateado: {archivo}")


def _crear_hoja_picking(ws, df, nombre_tienda, nombre_reporte, thin, medium, header_fill, alt_fill, control_fill):
    """Crea hoja completa de picking con header + datos."""
    
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    fila_actual = 1
    
    # ========== HEADER DE CONTROL ==========
    
    # FILA 1-2: TÍTULO
    ws.merge_cells('A1:F2')
    ws['A1'] = f"🏢 JAGI - {nombre_reporte.upper()}"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25
    
    # FILA 3: TIENDA Y FECHA
    ws.merge_cells('A3:C3')
    ws['A3'] = f"📍 TIENDA: {nombre_tienda}"
    ws['A3'].font = Font(bold=True, size=12, color="1F4E78")
    ws['A3'].alignment = Alignment(horizontal="left", vertical="center")
    ws['A3'].fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    
    ws.merge_cells('D3:F3')
    ws['D3'] = f"📅 Fecha: {fecha_actual}"
    ws['D3'].font = Font(bold=True, size=11)
    ws['D3'].alignment = Alignment(horizontal="right", vertical="center")
    ws['D3'].fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    ws.row_dimensions[3].height = 20
    
    # FILA 4: SEPARADOR
    ws.merge_cells('A4:F4')
    ws['A4'] = "━━━ CONTROL DE PICKING ━━━"
    ws['A4'].font = Font(bold=True, size=11, color="FFFFFF")
    ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A4'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.row_dimensions[4].height = 20
    
    # FILA 5: ESPACIO
    ws.row_dimensions[5].height = 5
    
    # FILA 6: HORA INICIO Y FINAL
    ws.merge_cells('A6:B6')
    ws['A6'] = "⏰ Hora Inicio:"
    ws['A6'].font = Font(bold=True, size=10)
    ws['A6'].alignment = Alignment(horizontal="right", vertical="center")
    ws['C6'].border = Border(bottom=thin)
    
    ws.merge_cells('D6:E6')
    ws['D6'] = "⏰ Hora Final:"
    ws['D6'].font = Font(bold=True, size=10)
    ws['D6'].alignment = Alignment(horizontal="right", vertical="center")
    ws['F6'].border = Border(bottom=thin)
    ws.row_dimensions[6].height = 20
    
    # FILA 7: ESPACIO
    ws.row_dimensions[7].height = 5
    
    # FILA 8: ENCARGADO
    ws.merge_cells('A8:B8')
    ws['A8'] = "👤 Encargado:"
    ws['A8'].font = Font(bold=True, size=10)
    ws['A8'].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells('C8:F8')
    # APLICAR EL BORDE A TODO EL RANGO (C8, D8, E8, F8)
    border_bottom = Border(bottom=thin)
    for row in ws['C8:F8']:
        for cell in row:
            cell.border = border_bottom
    ws.row_dimensions[8].height = 20
    
    # FILA 9: ESPACIO
    ws.row_dimensions[9].height = 10
    
    # ========== DATOS ==========
    
    fila_header = 10
    fila_datos_inicio = 11
    
    # Columna A: Checkbox
    ws.column_dimensions['A'].width = 4
    ws['A10'] = "☐"
    ws['A10'].font = Font(bold=True, size=11)
    ws['A10'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A10'].fill = header_fill
    ws['A10'].border = Border(top=medium, left=thin, right=thin, bottom=medium)
    
    # Headers de datos (columnas B en adelante)
    for col_idx, col_name in enumerate(df.columns, start=2):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=fila_header, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
        cell.border = Border(top=medium, left=thin, right=thin, bottom=medium)
    
    ws.row_dimensions[fila_header].height = 30
    
    # Escribir datos
    for i, (idx, row) in enumerate(df.iterrows()):
        fila = fila_datos_inicio + i
        
        # Checkbox
        ws.cell(row=fila, column=1).value = "☐"
        ws.cell(row=fila, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=fila, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        # Alternar color
        fill = alt_fill if fila % 2 == 0 else None
        if fill:
            ws.cell(row=fila, column=1).fill = fill
        
        # Datos
        for col_idx, value in enumerate(row, start=2):
            cell = ws.cell(row=fila, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if fill:
                cell.fill = fill
    
    # Auto-ajuste de columnas
    for col_idx in range(2, len(df.columns) + 2):
        col_letter = get_column_letter(col_idx)
        max_length = len(str(df.columns[col_idx - 2]))
        for cell in ws[col_letter]:
            if cell.value and cell.row >= fila_header:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 50)
    
    # Firma al final
    ultima_fila = fila_datos_inicio + len(df) + 2
    
    # Firma (Mezclamos A y B)
    ws.merge_cells(f'A{ultima_fila}:B{ultima_fila}')
    ws[f'A{ultima_fila}'] = "✍️ Firma:"
    ws[f'A{ultima_fila}'].font = Font(bold=True, size=11)
    ws[f'A{ultima_fila}'].alignment = Alignment(horizontal="right", vertical="center")

    # Columna C: Línea para firmar
    ws[f'C{ultima_fila}'].border = Border(bottom=thin)

    # Fecha (Mezclamos D y E)
    ws.merge_cells(f'D{ultima_fila}:E{ultima_fila}')
    ws[f'D{ultima_fila}'] = "📅 Fecha:"
    ws[f'D{ultima_fila}'].font = Font(bold=True, size=11)
    ws[f'D{ultima_fila}'].alignment = Alignment(horizontal="right", vertical="center")
    
    # Columna F: Línea para la fecha
    ws[f'F{ultima_fila}'].border = Border(bottom=thin)
    
    # Ajuste de altura para que el borde inferior se vea claro
    ws.row_dimensions[ultima_fila].height = 25
    
    # Configuración de impresión
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    
    ws.freeze_panes = f"A{fila_header + 1}"
    
    try:
        ws.oddHeader.left.text = f"&L{nombre_reporte}"
        ws.oddHeader.right.text = f"&R{nombre_tienda}"
        ws.oddFooter.center.text = "&CPágina &P de &N"
    except:
        pass


def _crear_hoja_general(ws, df, nombre_reporte, thin, header_fill, alt_fill):
    """Crea hoja estándar general."""
    
    # Escribir headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    ws.row_dimensions[1].height = 40
    
    # Escribir datos
    for i, (idx, row) in enumerate(df.iterrows()):
        fila = i + 2  # Fila 2 es la primera de datos (después del header en fila 1)
        fill = alt_fill if fila % 2 == 0 else None
        
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=fila, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if fill:
                cell.fill = fill
    
    # Auto-ajuste
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 50)
    
    ws.freeze_panes = "A2"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    
    try:
        ws.oddHeader.left.text = f"&LJAGI - {nombre_reporte}"
        ws.oddHeader.right.text = "&RPágina &P de &N"
        ws.oddFooter.center.text = "&CGenerado automáticamente"
    except:
        pass